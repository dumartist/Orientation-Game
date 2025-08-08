import os
import json
import uuid
from datetime import datetime
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'  # Change this in production

# User data storage (in production, use a proper database)
USERS_FILE = 'users.json'

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2)

class GameState:
    def __init__(self, user_id=None):
        self.player = {
            'name': 'Anomaly',
            'level': 1,
            'hp': 100,
            'max_hp': 100,
            'exp': 0,
            'exp_to_next': 100,
            'credits': 50,
            'inventory': [],
            'location': 'nexis',
            'quests': [],
            'skills': {
                'decryption': 1,
                'manipulation': 1,
                'reconstruction': 1
            },
            'reputation': {
                'codekeepers': 0,
                'resistance': 0,
                'neutral': 0
            }
        }
        self.game_log = []
        self.available_actions = []
        self.current_stage = 1
        self.story_progress = []
        self.save_id = None
        self.save_name = None
        self.save_date = None
        self.user_id = user_id

    def to_dict(self):
        return {
            'player': self.player,
            'game_log': self.game_log,
            'available_actions': self.available_actions,
            'current_stage': self.current_stage,
            'story_progress': self.story_progress,
            'save_id': self.save_id,
            'save_name': self.save_name,
            'save_date': self.save_date,
            'user_id': self.user_id
        }

    @classmethod
    def from_dict(cls, data, user_id=None):
        game_state = cls(user_id)
        game_state.player = data.get('player', game_state.player)
        game_state.game_log = data.get('game_log', [])
        game_state.available_actions = data.get('available_actions', [])
        game_state.current_stage = data.get('current_stage', 1)
        game_state.story_progress = data.get('story_progress', [])
        game_state.save_id = data.get('save_id')
        game_state.save_name = data.get('save_name')
        game_state.save_date = data.get('save_date')
        game_state.user_id = data.get('user_id', user_id)
        return game_state

# Initialize game state
game_state = GameState()

# Add initial game log entries
game_state.game_log = [
    {
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'message': 'Welcome to The Codebound Chronicles!',
        'type': 'story'
    },
    {
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'message': 'You are an Anomaly in the year 2125. The world is a digital simulation called the Source.',
        'type': 'story'
    },
    {
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'message': 'You have been brought to Nexis, the hidden heart of the Source, to train as a Codekeeper.',
        'type': 'story'
    },
    {
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'message': 'Your mentor Lira approaches. "Welcome, Anomaly. You can see the seams, can\'t you?"',
        'type': 'story'
    }
]

# Create saves directory if it doesn't exist
SAVE_DIR = 'saves'
if not os.path.exists(SAVE_DIR):
    os.makedirs(SAVE_DIR)

def get_save_file_path(save_id):
    return os.path.join(SAVE_DIR, f'{save_id}.json')

def save_game_to_file(game_state, save_name):
    save_id = str(uuid.uuid4())
    game_state.save_id = save_id
    game_state.save_name = save_name
    game_state.save_date = datetime.now().isoformat()
    
    save_data = game_state.to_dict()
    save_file_path = get_save_file_path(save_id)
    
    with open(save_file_path, 'w') as f:
        json.dump(save_data, f, indent=2)
    
    return save_id

def load_game_from_file(save_id, user_id=None):
    save_file_path = get_save_file_path(save_id)
    
    if not os.path.exists(save_file_path):
        return None
    
    with open(save_file_path, 'r') as f:
        save_data = json.load(f)
    
    # Check if user owns this save
    if user_id and save_data.get('user_id') != user_id:
        return None
    
    return GameState.from_dict(save_data, user_id)

def get_all_saves(user_id=None):
    saves = []
    
    if not os.path.exists(SAVE_DIR):
        return saves
    
    for filename in os.listdir(SAVE_DIR):
        if filename.endswith('.json'):
            save_id = filename[:-5]  # Remove .json extension
            save_file_path = os.path.join(SAVE_DIR, filename)
            
            try:
                with open(save_file_path, 'r') as f:
                    save_data = json.load(f)
                
                # Only include saves owned by the user
                if user_id and save_data.get('user_id') != user_id:
                    continue
                
                saves.append({
                    'save_id': save_id,
                    'save_name': save_data.get('save_name', 'Unknown Save'),
                    'save_date': save_data.get('save_date', ''),
                    'player_level': save_data.get('player', {}).get('level', 1),
                    'current_stage': save_data.get('current_stage', 1),
                    'user_id': save_data.get('user_id')
                })
            except:
                continue
    
    # Sort by save date (newest first)
    saves.sort(key=lambda x: x['save_date'], reverse=True)
    return saves

def delete_save_file(save_id, user_id=None):
    save_file_path = get_save_file_path(save_id)
    
    if not os.path.exists(save_file_path):
        return False
    
    # Check if user owns this save
    if user_id:
        try:
            with open(save_file_path, 'r') as f:
                save_data = json.load(f)
            if save_data.get('user_id') != user_id:
                return False
        except:
            return False
    
    os.remove(save_file_path)
    return True

@app.route('/')
def index():
    if 'user_id' not in session:
        return render_template('login.html')
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        npm = data.get('npm')
        
        users = load_users()
        
        if npm in users:
            session['user_id'] = npm
            session['username'] = users[npm]['username']
            return jsonify({'success': True, 'message': 'Login successful'})
        else:
            return jsonify({'success': False, 'message': 'NPM not found. Please register first.'})
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        npm = data.get('npm')
        
        if not username or not npm:
            return jsonify({'success': False, 'message': 'Username and NPM are required'})
        
        users = load_users()
        
        if npm in users:
            return jsonify({'success': False, 'message': 'NPM already registered'})
        
        # Store user data (in production, hash the NPM)
        users[npm] = {
            'username': username,
            'npm': npm,
            'created_at': datetime.now().isoformat()
        }
        
        save_users(users)
        
        return jsonify({'success': True, 'message': 'Registration successful. You can now login.'})
    
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/game-state')
def get_game_state():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Return only the last 20 log entries
    return jsonify({
        'player': game_state.player,
        'game_log': game_state.game_log[-20:],
        'available_actions': game_state.available_actions,
        'current_stage': game_state.current_stage,
        'story_progress': game_state.story_progress
    })

@app.route('/api/action', methods=['POST'])
def perform_action():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    action = data.get('action')
    target = data.get('target', '')
    
    if action == 'story_choice':
        result = make_story_choice(target)
        return jsonify(result)
    
    return jsonify({'success': False, 'message': 'Invalid action'})

def make_story_choice(choice):
    """Handle story choices and update game state"""
    current_stage = game_state.current_stage
    
    # Story outcomes based on stage and choice
    story_outcomes = {
        1: {  # The Awakening
            'a': {
                'message': 'You choose to trust Lira. She smiles warmly. "Good. Trust is the foundation of our work."',
                'credits_gain': 10,
                'skill_focus': 'decryption',
                'reputation_change': {'codekeepers': 2}
            },
            'b': {
                'message': 'You question everything. Lira\'s expression becomes thoughtful. "Curiosity can be dangerous, but also enlightening."',
                'credits_gain': 15,
                'skill_focus': 'manipulation',
                'reputation_change': {'resistance': 2}
            },
            'c': {
                'message': 'You focus on learning the three pillars. Lira nods approvingly. "Knowledge is power."',
                'credits_gain': 20,
                'skill_focus': 'reconstruction',
                'reputation_change': {'neutral': 2}
            },
            'd': {
                'message': 'You investigate the whispers. Lira\'s eyes narrow slightly. "You see more than most."',
                'credits_gain': 25,
                'skill_focus': 'decryption',
                'reputation_change': {'resistance': 1, 'codekeepers': 1}
            }
        },
        2: {  # The Trials Begin
            'a': {
                'message': 'You focus on personal advancement. Your skills grow rapidly.',
                'credits_gain': 30,
                'skill_focus': 'manipulation',
                'reputation_change': {'codekeepers': 3}
            },
            'b': {
                'message': 'You build alliances with questioning recruits. You gain trusted allies.',
                'credits_gain': 25,
                'skill_focus': 'reconstruction',
                'reputation_change': {'resistance': 3}
            },
            'c': {
                'message': 'You use your skills to help others. Your reputation grows.',
                'credits_gain': 20,
                'skill_focus': 'reconstruction',
                'reputation_change': {'neutral': 3}
            },
            'd': {
                'message': 'You sabotage the trials subtly. Your manipulation skills improve.',
                'credits_gain': 35,
                'skill_focus': 'manipulation',
                'reputation_change': {'resistance': 2, 'codekeepers': -1}
            }
        },
        3: {  # The Revelation
            'a': {
                'message': 'You embrace the power of the Grand Code. The system responds to your will.',
                'credits_gain': 50,
                'skill_focus': 'manipulation',
                'reputation_change': {'codekeepers': 5}
            },
            'b': {
                'message': 'You expose the truth to all humanity. The Source begins to fracture.',
                'credits_gain': 40,
                'skill_focus': 'decryption',
                'reputation_change': {'resistance': 5}
            },
            'c': {
                'message': 'You seek a third path. You begin subtle reforms to the Grand Code.',
                'credits_gain': 35,
                'skill_focus': 'reconstruction',
                'reputation_change': {'neutral': 5}
            },
            'd': {
                'message': 'You destroy the Grand Code entirely. Chaos erupts across the Source.',
                'credits_gain': 60,
                'skill_focus': 'decryption',
                'reputation_change': {'resistance': 4, 'codekeepers': -3}
            }
        },
        4: {  # The Final Choice
            'a': {
                'message': 'You choose The Collapse. Freedom is restored, but the perfect world fractures. Humanity must rebuild.',
                'credits_gain': 100,
                'ending': 'The Collapse'
            },
            'b': {
                'message': 'You choose The New Order. You become the new leader of The Codekeepers, reshaping reality.',
                'credits_gain': 100,
                'ending': 'The New Order'
            },
            'c': {
                'message': 'You choose The Balance. You find a delicate equilibrium between order and chaos.',
                'credits_gain': 100,
                'ending': 'The Balance'
            },
            'd': {
                'message': 'You choose The Anomaly. You reject all systems and forge your own path.',
                'credits_gain': 100,
                'ending': 'The Anomaly'
            }
        }
    }
    
    if current_stage not in story_outcomes or choice not in story_outcomes[current_stage]:
        return {'success': False, 'message': 'Invalid choice'}
    
    outcome = story_outcomes[current_stage][choice]
    
    # Apply outcome effects
    game_state.player['credits'] += outcome.get('credits_gain', 0)
    
    # Gain EXP from story choices
    exp_gain = outcome.get('credits_gain', 0) // 2  # EXP is half of credits gained
    game_state.player['exp'] += exp_gain
    
    # Check for level up
    while game_state.player['exp'] >= game_state.player['exp_to_next']:
        game_state.player['exp'] -= game_state.player['exp_to_next']
        game_state.player['level'] += 1
        game_state.player['exp_to_next'] = game_state.player['level'] * 100  # Increase EXP needed for next level
        game_state.player['max_hp'] += 20  # Increase max HP on level up
        game_state.player['hp'] = game_state.player['max_hp']  # Restore HP on level up
        
        # Add level up message to game log
        game_state.game_log.append({
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'message': f'Level up! You are now level {game_state.player["level"]}.',
            'type': 'level_up'
        })
    
    if 'skill_focus' in outcome:
        skill = outcome['skill_focus']
        game_state.player['skills'][skill] += 1
    
    if 'reputation_change' in outcome:
        for faction, change in outcome['reputation_change'].items():
            game_state.player['reputation'][faction] += change
    
    # Add to game log
    game_state.game_log.append({
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'message': outcome['message'],
        'type': 'story'
    })
    
    # Progress to next stage or end game
    if current_stage < 4:
        game_state.current_stage += 1
        game_state.game_log.append({
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'message': f'You have progressed to Stage {game_state.current_stage}.',
            'type': 'story'
        })
    else:
        # Game completed
        game_state.current_stage = 'complete'
        game_state.game_log.append({
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'message': f'Game completed! Ending: {outcome.get("ending", "Unknown")}',
            'type': 'story'
        })
    
    return {'success': True, 'message': outcome['message']}

@app.route('/api/update-actions')
def update_actions():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Update available actions based on current game state
    game_state.available_actions = ['story_choice']
    
    return jsonify({'available_actions': game_state.available_actions})

@app.route('/api/save-game', methods=['POST'])
def save_game():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    save_name = data.get('save_name', f'Game Save {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    
    try:
        # Associate save with current user
        game_state.user_id = session['user_id']
        save_id = save_game_to_file(game_state, save_name)
        return jsonify({'success': True, 'message': f'Game saved as "{save_name}"'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to save game: {str(e)}'})

@app.route('/api/load-game', methods=['POST'])
def load_game():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    save_id = data.get('save_id')
    
    try:
        loaded_state = load_game_from_file(save_id, session['user_id'])
        if loaded_state:
            global game_state
            game_state = loaded_state
            return jsonify({'success': True, 'message': 'Game loaded successfully', 'game_state': game_state.to_dict()})
        else:
            return jsonify({'success': False, 'message': 'Save file not found or access denied'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to load game: {str(e)}'})

@app.route('/api/list-saves')
def list_saves():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        saves = get_all_saves(session['user_id'])
        return jsonify({'success': True, 'saves': saves})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to list saves: {str(e)}'})

@app.route('/api/delete-save', methods=['POST'])
def delete_save():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    save_id = data.get('save_id')
    
    try:
        if delete_save_file(save_id, session['user_id']):
            return jsonify({'success': True, 'message': 'Save file deleted successfully'})
        else:
            return jsonify({'success': False, 'message': 'Save file not found or access denied'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to delete save: {str(e)}'})

@app.route('/api/restart-game', methods=['POST'])
def restart_game():
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        global game_state
        # Create a new game state for the current user
        game_state = GameState(session['user_id'])
        
        # Add initial game log entries
        game_state.game_log = [
            {
                'timestamp': datetime.now().strftime('%H:%M:%S'),
                'message': 'Welcome to The Codebound Chronicles!',
                'type': 'story'
            },
            {
                'timestamp': datetime.now().strftime('%H:%M:%S'),
                'message': 'You are an Anomaly in the year 2125. The world is a digital simulation called the Source.',
                'type': 'story'
            },
            {
                'timestamp': datetime.now().strftime('%H:%M:%S'),
                'message': 'You have been brought to Nexis, the hidden heart of the Source, to train as a Codekeeper.',
                'type': 'story'
            },
            {
                'timestamp': datetime.now().strftime('%H:%M:%S'),
                'message': 'Your mentor Lira approaches. "Welcome, Anomaly. You can see the seams, can\'t you?"',
                'type': 'story'
            }
        ]
        
        return jsonify({
            'success': True, 
            'message': 'Game restarted successfully!',
            'game_state': game_state.to_dict()
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to restart game: {str(e)}'})

@app.route('/downloadlistofpeserta')
def download_participant_list():
    """Download Excel file with list of registered participants"""
    try:
        # Load users data
        users = load_users()
        
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Daftar Peserta"
        
        # Add headers
        ws['A1'] = 'No.'
        ws['B1'] = 'NPM'
        ws['C1'] = 'Nama'
        ws['D1'] = 'Tanggal Registrasi'
        
        # Style the headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        row_num = 2
        for npm, user_data in users.items():
            ws[f'A{row_num}'] = row_num - 1  # No.
            ws[f'B{row_num}'] = npm  # NPM
            ws[f'C{row_num}'] = user_data.get('username', 'Unknown')  # Nama
            
            # Format the date to be human readable
            created_at = user_data.get('created_at', 'Unknown')
            if created_at != 'Unknown':
                try:
                    # Parse ISO format and convert to human readable
                    from datetime import datetime
                    dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    formatted_date = dt.strftime('%d/%m/%Y %H:%M:%S')
                    ws[f'D{row_num}'] = formatted_date
                except:
                    ws[f'D{row_num}'] = created_at
            else:
                ws[f'D{row_num}'] = created_at
            
            row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with current timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"daftar_peserta_{timestamp}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate participant list: {str(e)}'}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
